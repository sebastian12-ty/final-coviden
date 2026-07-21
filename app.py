"""
╔══════════════════════════════════════════════════════════════╗
║   COVIDEN DENTAL — Sistema Completo v4.0 PREMIUM            ║
║   Scan IA : Google Gemini 1.5 Flash                         ║
║   Chatbot : Groq  llama-3.3-70b-versatile                   ║
║   Extras  : Citas, Gamificacion, PWA, Reportes              ║
╠══════════════════════════════════════════════════════════════╣
║  pip install flask google-generativeai groq fpdf2           ║
║              openpyxl Pillow Werkzeug                       ║
║  python app.py  ->  http://localhost:5000                    ║
╚══════════════════════════════════════════════════════════════╝
"""

from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, session, flash, send_file)
import sqlite3, json, os, base64, re, io, hashlib, warnings
from datetime import datetime, timedelta, date
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

ENV_FILE = os.path.join(os.path.dirname(__file__), ".env.local")


def load_local_env(override=False):
    if not os.path.exists(ENV_FILE):
        return
    with open(ENV_FILE, "r", encoding="utf-8-sig") as fh:
        for raw in fh:
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, val = line.split("=", 1)
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            if key and (override or key not in os.environ):
                os.environ[key] = val


def save_local_env(updates):
    current, order = {}, []
    if os.path.exists(ENV_FILE):
        with open(ENV_FILE, "r", encoding="utf-8-sig") as fh:
            for raw in fh:
                line = raw.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, val = line.split("=", 1)
                key = key.strip()
                if key:
                    current[key] = val.strip().strip('"').strip("'")
                    order.append(key)
    for key, val in updates.items():
        if val is not None and str(val).strip():
            current[key] = str(val).strip()
            if key not in order:
                order.append(key)
    lines = [
        "# Local secrets for COVIDEN Dental",
        "# Do not commit this file to GitHub.",
        "",
    ]
    for key in order:
        if current.get(key):
            lines.append(f"{key}={current[key]}")
    tmp = ENV_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")
    os.replace(tmp, ENV_FILE)


def mask_secret(value):
    if not value:
        return "No configurada"
    if len(value) <= 8:
        return "Configurada"
    return f"Configurada ({value[:4]}...{value[-4:]})"


load_local_env()
# ── GOOGLE GEMINI / GROQ AI ────────────────────────────────────
with warnings.catch_warnings():
    warnings.simplefilter("ignore", FutureWarning)
    import google.generativeai as genai
from groq import Groq

# Fallback directo: pon aqui tus keys si no quieres usar .env.local
DEFAULT_GEMINI_API_KEY = "AQ.Ab8RN6JcJpKt8HB-gUqu1uTCZzSMThJ6ozNWtmwyxf3BABtsGQ"
DEFAULT_GEMINI_MODEL = "gemini-1.5-flash"
DEFAULT_GROQ_API_KEY = "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"
DEFAULT_GROQ_MODEL = "gpt-oss-120b"

GEMINI_API_KEY = DEFAULT_GEMINI_API_KEY
GEMINI_MODEL = DEFAULT_GEMINI_MODEL
gemini_model = None
GROQ_API_KEY = DEFAULT_GROQ_API_KEY
GROQ_MODEL = DEFAULT_GROQ_MODEL
groq_client = None


def refresh_ai_clients():
    global GEMINI_API_KEY, GEMINI_MODEL, gemini_model
    global GROQ_API_KEY, GROQ_MODEL, groq_client
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip() or DEFAULT_GEMINI_API_KEY
    GEMINI_MODEL = os.getenv("GEMINI_MODEL", "").strip() or DEFAULT_GEMINI_MODEL
    GROQ_API_KEY = os.getenv("GROQ_API_KEY", "").strip() or DEFAULT_GROQ_API_KEY
    GROQ_MODEL = os.getenv("GROQ_MODEL", "").strip() or DEFAULT_GROQ_MODEL
    gemini_model = None
    groq_client = None
    if GEMINI_API_KEY:
        genai.configure(api_key=GEMINI_API_KEY)
        gemini_model = genai.GenerativeModel(GEMINI_MODEL)
    if GROQ_API_KEY:
        groq_client = Groq(api_key=GROQ_API_KEY)


refresh_ai_clients()
print("="*60)
print("[DEBUG] GEMINI_API_KEY:", "OK ("+GEMINI_API_KEY[:6]+"...)" if GEMINI_API_KEY and "PON_TU" not in GEMINI_API_KEY else "VACIA o PLACEHOLDER ->", repr(GEMINI_API_KEY[:20]))
print("[DEBUG] gemini_model objeto:", gemini_model)
print("[DEBUG] GROQ_API_KEY:", "OK ("+GROQ_API_KEY[:6]+"...)" if GROQ_API_KEY and "PON_TU" not in GROQ_API_KEY else "VACIA o PLACEHOLDER ->", repr(GROQ_API_KEY[:20]))
print("[DEBUG] groq_client objeto:", groq_client)
print("[DEBUG] .env.local existe?:", os.path.exists(ENV_FILE), "->", ENV_FILE)
print("="*60)
# ── FLASK ──────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = "coviden_v4_ultra_secret_2024"
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

ADMIN_USER = "admin"
ADMIN_PASS = "coviden2024"

# ══════════════════════════════════════════════════════════════
#  BASE DE DATOS
# ══════════════════════════════════════════════════════════════
DB = "coviden.db"

def get_db():
    db = sqlite3.connect(DB)
    db.execute("PRAGMA journal_mode=MEMORY")
    db.execute("PRAGMA foreign_keys=ON")
    db.row_factory = sqlite3.Row
    return db

def init_db():
    db = get_db()
    db.executescript("""
    CREATE TABLE IF NOT EXISTS pacientes (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre        TEXT NOT NULL,
        apellido      TEXT NOT NULL,
        email         TEXT UNIQUE NOT NULL,
        telefono      TEXT,
        dni           TEXT,
        fecha_nac     TEXT,
        sexo          TEXT,
        direccion     TEXT,
        alergias      TEXT,
        password_hash TEXT NOT NULL,
        foto_perfil   TEXT,
        puntos        INTEGER DEFAULT 0,
        nivel         TEXT DEFAULT 'Bronce',
        tema          TEXT DEFAULT 'light',
        ultimo_saludo_cumple TEXT,
        created_at    TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS escaneos (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        paciente_id    INTEGER REFERENCES pacientes(id),
        nombre_libre   TEXT,
        email_libre    TEXT,
        telefono_libre TEXT,
        foto_base64    TEXT,
        expresion      TEXT,
        puntuacion     INTEGER,
        nivel_urgencia TEXT,
        diagnostico    TEXT,
        tratamientos   TEXT,
        hallazgos      TEXT,
        mensaje        TEXT,
        created_at     TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS citas (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        paciente_id  INTEGER REFERENCES pacientes(id),
        nombre_libre TEXT,
        email_libre  TEXT,
        tel_libre    TEXT,
        servicio     TEXT NOT NULL,
        fecha        TEXT NOT NULL,
        hora         TEXT NOT NULL,
        motivo       TEXT,
        estado       TEXT DEFAULT 'pendiente',
        notas_doctor TEXT,
        created_at   TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS pagos (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        paciente_id  INTEGER REFERENCES pacientes(id),
        nombre_libre TEXT,
        email_libre  TEXT,
        tel_libre    TEXT,
        cita_id      INTEGER REFERENCES citas(id),
        servicio     TEXT NOT NULL,
        monto        REAL NOT NULL,
        metodo       TEXT NOT NULL,
        estado       TEXT DEFAULT 'pendiente',
        referencia   TEXT,
        captura_b64  TEXT,
        notas        TEXT,
        created_at   TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS chat_mensajes (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        paciente_id INTEGER REFERENCES pacientes(id),
        rol         TEXT NOT NULL,
        contenido   TEXT NOT NULL,
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS resenas (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        paciente_id INTEGER REFERENCES pacientes(id),
        nombre      TEXT NOT NULL,
        servicio    TEXT NOT NULL,
        estrellas   INTEGER NOT NULL CHECK(estrellas BETWEEN 1 AND 5),
        comentario  TEXT,
        util_si     INTEGER DEFAULT 0,
        util_no     INTEGER DEFAULT 0,
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS votos (
        resena_id   INTEGER,
        session_key TEXT,
        PRIMARY KEY(resena_id, session_key)
    );
    CREATE TABLE IF NOT EXISTS notificaciones (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        paciente_id INTEGER REFERENCES pacientes(id),
        tipo        TEXT,
        titulo      TEXT,
        mensaje     TEXT,
        leida       INTEGER DEFAULT 0,
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS inventario (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre      TEXT NOT NULL,
        categoria   TEXT NOT NULL,
        cantidad    INTEGER DEFAULT 0,
        unidad      TEXT,
        minimo      INTEGER DEFAULT 5,
        precio_unit REAL DEFAULT 0,
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS historial_clinico (
        id                     INTEGER PRIMARY KEY AUTOINCREMENT,
        paciente_id            INTEGER REFERENCES pacientes(id),
        fecha                  TEXT DEFAULT (date('now','localtime')),
        tipo                   TEXT,
        diagnostico            TEXT,
        tratamiento_realizado  TEXT,
        medicamentos           TEXT,
        odontograma            TEXT,
        observaciones          TEXT,
        proxima_cita           TEXT,
        creado_por             TEXT DEFAULT 'Dr. Sergio Castillo',
        created_at             TEXT DEFAULT (datetime('now','localtime'))
    );
    """)
    db.commit()
    db.close()

# ── AUTH ───────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def dec(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return dec

def login_required(f):
    @wraps(f)
    def dec(*args, **kwargs):
        if not session.get("pid"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return dec

def current_patient():
    pid = session.get("pid")
    if not pid: return None
    db = get_db()
    p = db.execute("SELECT * FROM pacientes WHERE id=?", (pid,)).fetchone()
    db.close()
    return p

def add_points(pid, pts, motivo="Acción"):
    """Añade puntos y actualiza nivel."""
    if not pid: return
    db = get_db()
    db.execute("UPDATE pacientes SET puntos=puntos+? WHERE id=?", (pts, pid))
    row = db.execute("SELECT puntos FROM pacientes WHERE id=?", (pid,)).fetchone()
    total = row["puntos"] if row else 0
    nivel = "Bronce" if total < 100 else "Plata" if total < 300 else "Oro" if total < 600 else "Diamante"
    db.execute("UPDATE pacientes SET nivel=? WHERE id=?", (nivel, pid))
    db.execute("INSERT INTO notificaciones (paciente_id,tipo,titulo,mensaje) VALUES (?,?,?,?)",
               (pid, "puntos", f"+{pts} puntos", f"Ganaste {pts} puntos por: {motivo}"))
    db.commit(); db.close()

def calcular_edad(fecha_nac):
    """Calcula la edad en años a partir de YYYY-MM-DD."""
    if not fecha_nac:
        return None
    try:
        y, m, d = [int(x) for x in fecha_nac.split("-")]
        hoy = date.today()
        return hoy.year - y - ((hoy.month, hoy.day) < (m, d))
    except Exception:
        return None

HORAS_ATENCION = {"09:00","10:00","11:00","12:00","14:00","15:00","16:00","17:00","18:00","19:00"}


def parse_json_body():
    return request.get_json(silent=True) or {}


def normalizar_hora(hora):
    hora = (hora or "").strip()
    if len(hora) == 4 and hora[1] == ":":
        hora = "0" + hora
    return hora


def fecha_valida(fecha, permitir_pasado=False):
    try:
        valor = datetime.strptime(fecha, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None
    if not permitir_pasado and valor < date.today():
        return None
    return valor

def revisar_cumpleanos():
    """Revisa cumpleaños del día y envía notificación + puntos de regalo
    (una sola vez por año por paciente)."""
    db = get_db()
    hoy = date.today()
    hoy_mmdd = hoy.strftime("%m-%d")
    hoy_iso  = hoy.isoformat()
    pacientes = db.execute(
        "SELECT id,nombre,fecha_nac,ultimo_saludo_cumple FROM pacientes WHERE fecha_nac IS NOT NULL AND fecha_nac != ''"
    ).fetchall()
    saludados = []
    for p in pacientes:
        try:
            _, mm, dd = p["fecha_nac"].split("-")
        except Exception:
            continue
        if f"{mm}-{dd}" != hoy_mmdd:
            continue
        if p["ultimo_saludo_cumple"] == hoy_iso:
            continue
        db.execute("""INSERT INTO notificaciones (paciente_id,tipo,titulo,mensaje) VALUES (?,?,?,?)""",
            (p["id"], "cumpleanos", f"🎉 ¡Feliz Cumpleaños, {p['nombre']}!",
             "Todo el equipo de COVIDEN Dental te desea un excelente día. ¡De regalo, te obsequiamos 25 puntos! 🎁🦷"))
        db.execute("UPDATE pacientes SET puntos=puntos+25, ultimo_saludo_cumple=? WHERE id=?",
                   (hoy_iso, p["id"]))
        row = db.execute("SELECT puntos FROM pacientes WHERE id=?",(p["id"],)).fetchone()
        total = row["puntos"] if row else 0
        nivel = "Bronce" if total<100 else "Plata" if total<300 else "Oro" if total<600 else "Diamante"
        db.execute("UPDATE pacientes SET nivel=? WHERE id=?", (nivel,p["id"]))
        saludados.append(p["nombre"])
    db.commit(); db.close()
    return saludados

# ══════════════════════════════════════════════════════════════
#  PROMPTS IA
# ══════════════════════════════════════════════════════════════
SCAN_PROMPT = """Eres un asistente dental especializado de COVIDEN Dental, Lima, Peru.
Analiza CUIDADOSAMENTE la imagen dental/boca proporcionada.

Responde UNICAMENTE con JSON valido, sin markdown ni texto extra:
{
  "es_dental": true,
  "expresion": "sonrisa amplia",
  "puntuacion": 7,
  "nivel_urgencia": "media",
  "hallazgos": ["Hallazgo 1", "Hallazgo 2"],
  "tratamientos": [
    {"nombre":"Blanqueamiento","descripcion":"Descripcion breve.","precio":"S/ 350","prioridad":"rutina"}
  ],
  "diagnostico": "Parrafo diagnostico profesional.",
  "mensaje": "Mensaje empatico y motivador.",
  "recomendaciones_casa": ["Cepilla 3 veces al dia","Usa hilo dental diariamente"]
}

Reglas:
- Si NO ves dientes/boca: es_dental: false, tratamientos: []
- puntuacion: 1(muy malo) a 10(excelente)
- nivel_urgencia: baja | media | alta | urgente
- prioridad: rutina | pronto | inmediata
- Maximo 4 tratamientos relevantes, precios en soles peruanos (S/)
- recomendaciones_casa: 3-4 tips practicos para el paciente
"""

CHAT_SYSTEM = """Eres el asistente virtual de COVIDEN Dental (Comas, Lima, Peru).
Nombre del asistente: Covi

DATOS CLINICOS:
- Direccion: Av. Universitaria 8156, Comas, Lima
- Tel: (01) 123-4567 | WhatsApp: 944 439 324 | Email: contacto@coviden.com
- Horario: Lun-Sab 9am-8pm (domingos cerrado)
- Dr. Sergio Castillo Quispe - Odontologo Principal (+20 anios exp.)
- RUC: 20554909702 | Fundada: Oct 2013

SERVICIOS Y PRECIOS COMPLETOS:
Preventivos:
- Consulta General / Evaluacion: S/ 40
- Limpieza Dental (profilaxis): S/ 80 (recomendada cada 6 meses)
- Radiografia periapical: S/ 30 | Panoramica: S/ 80
- Fluoruracion: S/ 50 | Sellantes: S/ 40 por pieza

Esteticos:
- Blanqueamiento dental (1 sesion): S/ 350
- Carillas de porcelana: S/ 400 por diente
- Resina estetica / empaste: S/ 80-150 por pieza
- Microabrasion del esmalte: S/ 120

Restauradores:
- Empaste resina compuesta: S/ 80-150
- Incrustaciones (inlay/onlay): S/ 300-500
- Corona metalica: S/ 400 | Porcelana: S/ 700 | Zirconia: S/ 1,200

Ortodoncia:
- Ortodoncia metalica: S/ 2,500
- Ortodoncia ceramica: S/ 3,200
- Ortodoncia invisible (alineadores): S/ 4,500
- Retenedor: S/ 200

Cirugia:
- Extraccion simple: S/ 80 | Muela de juicio: S/ 250-400
- Cirugia de encias (periodoncia): S/ 300-600
- Regeneracion osea: S/ 500+
- Frenectomia: S/ 300

Implantologia:
- Implante dental titanio: desde S/ 1,800
- Implante con corona: S/ 2,500-3,500
- Sobredentadura sobre implantes: S/ 4,000+

Protesis:
- Protesis removible parcial: S/ 600
- Protesis total: S/ 800
- Protesis fija (puente): S/ 1,200-2,000
- Protesis sobre implantes: S/ 2,500+

Rehabilitacion:
- Rehabilitacion oral completa: segun evaluacion
- Tratamiento de conductos (endodoncia): S/ 350-500
- Tratamiento periodontal: S/ 400-800

PAGOS: Yape (944 439 324), Plin, Transferencia Interbank, o efectivo.

GAMIFICACION:
- Registro: +50 pts | Escaneo: +20 pts | Cita: +30 pts | Resena: +15 pts | Pago confirmado: +10 pts | Cumpleanos: +25 pts
- Niveles: Bronce(0-99), Plata(100-299), Oro(300-599), Diamante(600+)

INSTRUCCIONES:
- Responde SIEMPRE en espanol, amable y profesional como "Covi"
- Respuestas concisas (max 4-5 oraciones)
- Usa emojis con moderacion
- Si el usuario tiene dolor o urgencia, recomienda llamar INMEDIATAMENTE al (01) 123-4567
- Puedes orientar sobre citas, precios, ubicacion, horarios, tratamientos
- Si no sabes el precio exacto, sugiere agendar una consulta de evaluacion (S/ 40)"""

# ══════════════════════════════════════════════════════════════
#  RUTAS PUBLICAS
# ══════════════════════════════════════════════════════════════
@app.route("/")
def index():
    revisar_cumpleanos()  # Verifica cumpleaños en cada visita pública
    db = get_db()
    resenas   = db.execute("SELECT * FROM resenas ORDER BY created_at DESC LIMIT 12").fetchall()
    avg       = db.execute("SELECT AVG(estrellas) as a FROM resenas").fetchone()["a"]
    total_res = db.execute("SELECT COUNT(*) as c FROM resenas").fetchone()["c"]
    t_pac  = db.execute("SELECT COUNT(*) as c FROM pacientes").fetchone()["c"]
    t_scan = db.execute("SELECT COUNT(*) as c FROM escaneos").fetchone()["c"]
    t_cita = db.execute("SELECT COUNT(*) as c FROM citas WHERE estado='confirmada'").fetchone()["c"]
    db.close()
    return render_template("index.html",
        resenas=resenas, avg_rating=round(avg,1) if avg else 0,
        total_resenas=total_res, paciente=current_patient(),
        t_pac=t_pac, t_scan=t_scan, t_cita=t_cita)

# ── AUTH PACIENTE ──────────────────────────────────────────────
@app.route("/registro", methods=["GET","POST"])
def registro():
    if request.method == "POST":
        nombre = request.form.get("nombre","").strip()
        apellido = request.form.get("apellido","").strip()
        email = request.form.get("email","").strip().lower()
        tel   = request.form.get("telefono","").strip()
        pwd   = request.form.get("password","")
        pwd2  = request.form.get("password2","")
        if not nombre or not apellido or not email or not pwd:
            flash("Completa todos los campos obligatorios.", "err")
            return render_template("registro.html")
        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            flash("Ingresa un correo valido.", "err")
            return render_template("registro.html")
        if len(pwd) < 8:
            flash("La contrasena debe tener al menos 8 caracteres.", "err")
            return render_template("registro.html")
        if pwd2 and pwd != pwd2:
            flash("Las contrasenas no coinciden.", "err")
            return render_template("registro.html")
        db = get_db()
        if db.execute("SELECT id FROM pacientes WHERE email=?",(email,)).fetchone():
            db.close(); flash("Ese correo ya esta registrado.","err")
            return render_template("registro.html")
        cur = db.execute(
            "INSERT INTO pacientes (nombre,apellido,email,telefono,password_hash,puntos) VALUES (?,?,?,?,?,?)",
            (nombre,apellido,email,tel,generate_password_hash(pwd),50))
        pid = cur.lastrowid
        db.execute("INSERT INTO notificaciones (paciente_id,tipo,titulo,mensaje) VALUES (?,?,?,?)",
                   (pid,"bienvenida","?Bienvenido a COVIDEN!","Obtuviste 50 puntos de bienvenida ??"))
        db.commit(); db.close()
        session["pid"] = pid
        session["pnombre"] = f"{nombre} {apellido}"
        return redirect(url_for("dashboard"))
    return render_template("registro.html")

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email","").strip()
        pwd   = request.form.get("password","")
        db = get_db()
        p = db.execute("SELECT * FROM pacientes WHERE email=?",(email,)).fetchone()
        db.close()
        if p and check_password_hash(p["password_hash"], pwd):
            session["pid"] = p["id"]
            session["pnombre"] = f"{p['nombre']} {p['apellido']}"
            return redirect(url_for("dashboard"))
        flash("Credenciales incorrectas.","err")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))

# ── DASHBOARD PACIENTE ─────────────────────────────────────────
@app.route("/mi-cuenta")
@login_required
def dashboard():
    pid = session["pid"]
    revisar_cumpleanos()
    db  = get_db()
    pac       = db.execute("SELECT * FROM pacientes WHERE id=?", (pid,)).fetchone()
    escaneos  = db.execute("SELECT * FROM escaneos WHERE paciente_id=? ORDER BY created_at DESC LIMIT 20", (pid,)).fetchall()
    citas     = db.execute("SELECT * FROM citas WHERE paciente_id=? ORDER BY fecha DESC, hora DESC LIMIT 20", (pid,)).fetchall()
    pagos     = db.execute("SELECT * FROM pagos WHERE paciente_id=? ORDER BY created_at DESC LIMIT 20", (pid,)).fetchall()
    chats     = db.execute("SELECT * FROM chat_mensajes WHERE paciente_id=? ORDER BY created_at DESC LIMIT 40", (pid,)).fetchall()
    notifs    = db.execute("SELECT * FROM notificaciones WHERE paciente_id=? AND leida=0 ORDER BY created_at DESC LIMIT 30", (pid,)).fetchall()
    historial = db.execute("SELECT * FROM historial_clinico WHERE paciente_id=? ORDER BY fecha DESC, created_at DESC", (pid,)).fetchall()
    db.close()
    edad = calcular_edad(pac["fecha_nac"]) if pac else None
    return render_template("dashboard_paciente.html",
        paciente=pac, escaneos=escaneos, citas=citas,
        pagos=pagos, chats=chats, notifs=notifs,
        historial=historial, edad=edad)

@app.route("/mi-cuenta/notificaciones")
@login_required
def get_notificaciones():
    pid = session["pid"]
    db  = get_db()
    rows = db.execute(
        "SELECT * FROM notificaciones WHERE paciente_id=? AND leida=0 ORDER BY created_at DESC LIMIT 30",
        (pid,)).fetchall()
    db.close()
    return jsonify({"ok":True,"count":len(rows),"notifs":[dict(r) for r in rows]})

@app.route("/mi-cuenta/foto", methods=["POST"])
@login_required
def update_foto():
    data = request.get_json()
    foto = data.get("foto","")
    if "," in foto: foto = foto.split(",")[1]
    db = get_db()
    db.execute("UPDATE pacientes SET foto_perfil=? WHERE id=?", (foto[:300000], session["pid"]))
    db.commit(); db.close()
    return jsonify({"ok":True})

@app.route("/mi-cuenta/perfil", methods=["POST"])
@login_required
def update_perfil():
    d = request.get_json()
    db = get_db()
    db.execute("""UPDATE pacientes
        SET nombre=?,apellido=?,telefono=?,dni=?,fecha_nac=?,sexo=?,direccion=?,alergias=?,tema=?
        WHERE id=?""",
               (d.get("nombre"), d.get("apellido"), d.get("telefono"),
                d.get("dni"), d.get("fecha_nac"), d.get("sexo",""),
                d.get("direccion",""), d.get("alergias",""),
                d.get("tema","light"), session["pid"]))
    db.commit(); db.close()
    session["pnombre"] = f"{d.get('nombre','')} {d.get('apellido','')}"
    return jsonify({"ok":True})

@app.route("/mi-cuenta/marcar-notif", methods=["POST"])
@login_required
def marcar_notif():
    d   = request.get_json(silent=True) or {}
    nid = d.get("id")
    db  = get_db()
    if nid:
        db.execute("UPDATE notificaciones SET leida=1 WHERE id=? AND paciente_id=?", (nid, session["pid"]))
    else:
        db.execute("UPDATE notificaciones SET leida=1 WHERE paciente_id=?", (session["pid"],))
    db.commit(); db.close()
    return jsonify({"ok":True})



# ── ESCANEO DENTAL ─────────────────────────────────────────────
@app.route("/analizar", methods=["POST"])
def analizar():
    try:
        data    = parse_json_body()
        img_b64 = data.get("imagen","")
        if not img_b64:
            return jsonify({"ok":False,"msg":"Envia una imagen para analizar."}), 400
        if "," in img_b64: img_b64 = img_b64.split(",")[1]
        img_bytes = base64.b64decode(img_b64)
        nombre  = data.get("nombre","Paciente").strip()
        email   = data.get("email","").strip()
        tel     = data.get("telefono","").strip()
        pid     = session.get("pid")

        # ── Gemini API ─────────────────────────────────────────
        if gemini_model is None:
            return jsonify({"ok":False,"msg":"Configura GEMINI_API_KEY para activar el analisis IA."}), 503

        response = gemini_model.generate_content(
            [SCAN_PROMPT, {"mime_type":"image/jpeg","data":img_bytes}],
            request_options={"timeout": 60}
        )
        raw    = response.text.strip()
        raw    = re.sub(r"^```json\s*|```$","",raw,flags=re.MULTILINE).strip()
        result = json.loads(raw)

        if not result.get("es_dental",True):
            return jsonify({"ok":False,"msg":"No detectamos dientes en la imagen. Acerca la cámara a tu boca y sonríe. 😁"})

        foto = img_b64[:280000] if len(img_b64)>280000 else img_b64
        db   = get_db()
        cur  = db.execute("""INSERT INTO escaneos
            (paciente_id,nombre_libre,email_libre,telefono_libre,foto_base64,
             expresion,puntuacion,nivel_urgencia,diagnostico,tratamientos,hallazgos,mensaje)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (pid,nombre,email,tel,foto,
             result.get("expresion","neutral"), result.get("puntuacion",5),
             result.get("nivel_urgencia","media"), result.get("diagnostico",""),
             json.dumps(result.get("tratamientos",[]),ensure_ascii=False),
             json.dumps(result.get("hallazgos",[]),ensure_ascii=False),
             result.get("mensaje","")))
        scan_id = cur.lastrowid
        db.commit(); db.close()

        if pid: add_points(pid, 20, "Escaneo dental completado")
        return jsonify({"ok":True,"scan_id":scan_id,"resultado":result})

    except json.JSONDecodeError:
        return jsonify({"ok":False,"msg":"Error procesando análisis IA. Intenta de nuevo."})
    except Exception as e:
        print(f"[Gemini] {e}")
        return jsonify({"ok":False,"msg":str(e)})

# ── REPORTE PDF DEL ESCANEO ────────────────────────────────────
@app.route("/escaneo/<int:sid>/pdf")
def escaneo_pdf(sid):
    from fpdf import FPDF
    db = get_db()
    e  = db.execute("SELECT e.*,p.nombre as pn,p.apellido as pa FROM escaneos e LEFT JOIN pacientes p ON e.paciente_id=p.id WHERE e.id=?",(sid,)).fetchone()
    db.close()
    if not e: return "No encontrado",404

    pdf = FPDF()
    pdf.core_fonts_encoding = "cp1252"
    pdf.add_page()
    pdf.set_fill_color(5,14,39)
    pdf.rect(0,0,210,40,'F')
    pdf.set_font("Helvetica","B",18)
    pdf.set_text_color(255,255,255)
    pdf.set_xy(10,10)
    pdf.cell(0,10,"COVIDEN DENTAL - Reporte de Analisis IA",ln=True)
    pdf.set_font("Helvetica","",10)
    pdf.set_xy(10,24)
    pdf.cell(0,6,f"Av. Universitaria 8156, Comas, Lima | Tel: (01) 123-4567",ln=True)
    pdf.set_text_color(0,0,0)
    pdf.ln(12)

    nombre = f"{e['pn'] or ''} {e['pa'] or ''}".strip() or e["nombre_libre"] or "Anonimo"
    pdf.set_font("Helvetica","B",13)
    pdf.cell(0,8,f"Paciente: {nombre}",ln=True)
    pdf.set_font("Helvetica","",10)
    pdf.cell(0,6,f"Fecha: {e['created_at'][:16]}  |  Puntuacion: {e['puntuacion'] or 5}/10  |  Urgencia: {(e['nivel_urgencia'] or 'media').upper()}",ln=True)
    pdf.cell(0,6,f"Expresion detectada: {e['expresion'] or 'neutral'}",ln=True)
    pdf.ln(4)

    if e["diagnostico"]:
        pdf.set_font("Helvetica","B",11); pdf.cell(0,7,"DIAGNOSTICO:",ln=True)
        pdf.set_font("Helvetica","",9)
        pdf.multi_cell(0,5,e["diagnostico"])
        pdf.ln(3)

    hallazgos = json.loads(e["hallazgos"] or "[]")
    if hallazgos:
        pdf.set_font("Helvetica","B",11); pdf.cell(0,7,"HALLAZGOS DETECTADOS:",ln=True)
        pdf.set_font("Helvetica","",9)
        for h in hallazgos: pdf.cell(0,5,f"  > {h}",ln=True)
        pdf.ln(3)

    trts = json.loads(e["tratamientos"] or "[]")
    if trts:
        pdf.set_font("Helvetica","B",11); pdf.cell(0,7,"TRATAMIENTOS RECOMENDADOS:",ln=True)
        pdf.set_font("Helvetica","B",9)
        cols = [("Tratamiento",80),("Precio",30),("Prioridad",35),("Descripcion",55)]
        for c,w in cols: pdf.cell(w,7,c,border=1,align="C",fill=True)
        pdf.set_fill_color(255,255,255); pdf.ln()
        pdf.set_font("Helvetica","",8)
        for t in trts:
            pdf.cell(80,6,(t.get("nombre",""))[:35],border=1)
            pdf.cell(30,6,t.get("precio",""),border=1,align="C")
            pdf.cell(35,6,(t.get("prioridad","")).upper(),border=1,align="C")
            pdf.cell(55,6,(t.get("descripcion",""))[:28],border=1)
            pdf.ln()
        pdf.ln(4)

    if e["mensaje"]:
        pdf.set_fill_color(240,246,255)
        pdf.set_font("Helvetica","I",9)
        pdf.multi_cell(0,5,f'"{e["mensaje"]}"',fill=True)
        pdf.ln(2)

    pdf.set_font("Helvetica","",8)
    pdf.set_text_color(100,116,139)
    pdf.cell(0,5,"* Este reporte es un diagnostico preliminar generado por IA. No reemplaza una consulta medica profesional.",ln=True)
    pdf.cell(0,5,f"Generado por COVIDEN Dental - Modelo: {GEMINI_MODEL}",ln=True)

    buf = io.BytesIO(pdf.output()); buf.seek(0)
    return send_file(buf,mimetype="application/pdf",as_attachment=True,
        download_name=f"coviden_diagnostico_{sid}.pdf")

# ── CITAS ──────────────────────────────────────────────────────
@app.route("/agendar-cita", methods=["POST"])
def agendar_cita():
    d        = parse_json_body()
    pid      = session.get("pid")
    nombre   = d.get("nombre","").strip()
    email    = d.get("email","").strip().lower()
    tel      = d.get("telefono","").strip()
    servicio = d.get("servicio","").strip()
    fecha    = d.get("fecha","").strip()
    hora     = normalizar_hora(d.get("hora","").strip())
    motivo   = d.get("motivo","").strip()

    if not nombre and not pid:
        return jsonify({"ok":False,"msg":"Ingresa tu nombre."}), 400
    if not all([servicio, fecha, hora]):
        return jsonify({"ok":False,"msg":"Completa todos los campos."}), 400
    if not fecha_valida(fecha):
        return jsonify({"ok":False,"msg":"Elige una fecha valida desde hoy."}), 400
    if hora not in HORAS_ATENCION:
        return jsonify({"ok":False,"msg":"Elige un horario disponible."}), 400

    db = get_db()
    conflict = db.execute(
        "SELECT id FROM citas WHERE fecha=? AND hora=? AND estado NOT IN ('cancelada','rechazada')",
        (fecha, hora)).fetchone()
    if conflict:
        db.close()
        return jsonify({"ok":False,"conflict":True,
            "msg":f"El horario {hora} del {fecha} ya fue reservado. Elige otro horario."})

    cur = db.execute("""INSERT INTO citas
        (paciente_id,nombre_libre,email_libre,tel_libre,servicio,fecha,hora,motivo)
        VALUES (?,?,?,?,?,?,?,?)""",
        (pid,nombre,email,tel,servicio,fecha,hora,motivo))
    cita_id = cur.lastrowid
    if pid:
        db.execute("INSERT INTO notificaciones (paciente_id,tipo,titulo,mensaje) VALUES (?,?,?,?)",
                   (pid,"cita","Cita agendada",
                    f"Tu cita de {servicio} el {fecha} a las {hora} fue registrada. ?"))
    db.commit(); db.close()
    if pid: add_points(pid, 30, "Cita agendada")
    return jsonify({"ok":True,"cita_id":cita_id,
        "msg":f"Cita #{cita_id} confirmada para el {fecha} a las {hora}."})

@app.route("/disponibilidad")
def disponibilidad():
    fecha = request.args.get("fecha","")
    if not fecha or not fecha_valida(fecha): return jsonify({"ocupadas":[]})
    db = get_db()
    rows = db.execute(
        "SELECT hora FROM citas WHERE fecha=? AND estado NOT IN ('cancelada','rechazada')",
        (fecha,)).fetchall()
    db.close()
    ocupadas = []
    for r in rows:
        h = (r["hora"] or "").strip()
        if len(h)==4 and h[1]==":": h = "0"+h  # "9:00" -> "09:00"
        ocupadas.append(h)
    return jsonify({"ocupadas": ocupadas})

# ── PAGOS ──────────────────────────────────────────────────────
@app.route("/pago", methods=["POST"])
def crear_pago():
    d        = parse_json_body()
    pid      = session.get("pid")
    servicio = d.get("servicio","").strip()
    try:
        monto = float(d.get("monto",0) or 0)
    except (TypeError, ValueError):
        monto = 0
    metodo   = d.get("metodo","").strip()
    ref      = d.get("referencia","").strip()
    notas    = d.get("notas","").strip()
    nombre   = d.get("nombre","").strip()
    email    = d.get("email","").strip().lower()
    tel      = d.get("telefono","").strip()
    captura  = d.get("captura","")
    if "," in captura: captura = captura.split(",")[1]

    if not servicio or not metodo or monto<=0:
        return jsonify({"ok":False,"msg":"Completa todos los campos."}), 400
    if metodo not in {"yape","plin","transferencia","efectivo"}:
        return jsonify({"ok":False,"msg":"Metodo de pago no valido."}), 400

    db = get_db()
    if pid and not email:
        prow = db.execute("SELECT email,telefono FROM pacientes WHERE id=?",(pid,)).fetchone()
        if prow:
            email = email or (prow["email"] or "")
            tel   = tel   or (prow["telefono"] or "")

    cur = db.execute("""INSERT INTO pagos
        (paciente_id,nombre_libre,email_libre,tel_libre,servicio,monto,metodo,referencia,captura_b64,notas)
        VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (pid,nombre,email,tel,servicio,monto,metodo,ref,
         captura[:200000] if captura else "",notas))
    pago_id = cur.lastrowid
    if pid:
        db.execute("INSERT INTO notificaciones (paciente_id,tipo,titulo,mensaje) VALUES (?,?,?,?)",
                   (pid,"pago","Pago registrado",
                    f"Tu pago de S/ {monto:.2f} por {servicio} est? pendiente de confirmaci?n."))
    db.commit(); db.close()
    return jsonify({"ok":True,"pago_id":pago_id,
        "msg":f"Pago #{pago_id} registrado. Confirmaremos pronto."})

@app.route("/chat", methods=["POST"])
def chat():
    d    = parse_json_body()
    msgs = d.get("messages",[])
    if not isinstance(msgs, list):
        return jsonify({"ok":False,"reply":"Mensaje invalido."}), 400
    pid  = session.get("pid")
    if groq_client is None:
        return jsonify({"ok":False,"reply":"Configura GROQ_API_KEY para activar el asistente IA."}), 503
    try:
        completion = groq_client.chat.completions.create(
            model=GROQ_MODEL,
            messages=[{"role":"system","content":CHAT_SYSTEM}] +
                     [{"role":m["role"],"content":m["content"]} for m in msgs],
            temperature=0.72, max_tokens=700, stream=False
        )
        reply = completion.choices[0].message.content.strip()
        if pid and msgs:
            db = get_db()
            db.execute("INSERT INTO chat_mensajes (paciente_id,rol,contenido) VALUES (?,?,?)",
                       (pid,"user",msgs[-1]["content"]))
            db.execute("INSERT INTO chat_mensajes (paciente_id,rol,contenido) VALUES (?,?,?)",
                       (pid,"assistant",reply))
            db.commit(); db.close()
        return jsonify({"ok":True,"reply":reply})
    except Exception as e:
        print(f"[Groq] {e}")
        return jsonify({"ok":False,"reply":f"Error: {str(e)}"})

# ── RESEÑAS ────────────────────────────────────────────────────
@app.route("/resena", methods=["POST"])
def crear_resena():
    d = parse_json_body()
    pid = session.get("pid")
    nombre = d.get("nombre","").strip()
    servicio = d.get("servicio","").strip()
    try:
        estrellas = int(d.get("estrellas",0))
    except (TypeError, ValueError):
        estrellas = 0
    comentario = d.get("comentario","").strip()
    if not nombre or not servicio or not (1<=estrellas<=5):
        return jsonify({"ok":False,"msg":"Completa todos los campos."}), 400
    db = get_db()
    db.execute("INSERT INTO resenas (paciente_id,nombre,servicio,estrellas,comentario) VALUES (?,?,?,?,?)",
               (pid,nombre,servicio,estrellas,comentario))
    db.commit(); db.close()
    if pid: add_points(pid, 15, "Reseña publicada")
    return jsonify({"ok":True,"msg":"¡Gracias por tu reseña! 🌟"})

@app.route("/votar", methods=["POST"])
def votar():
    d = parse_json_body()
    try:
        rid = int(d.get("id",0))
    except (TypeError, ValueError):
        rid = 0
    voto = d.get("voto")
    if rid <= 0 or voto not in ["si","no"]:
        return jsonify({"ok":False,"msg":"Voto invalido."}), 400

    sk  = session.get("sk")
    if not sk:
        import uuid; sk = str(uuid.uuid4()); session["sk"] = sk
    db = get_db()
    if not db.execute("SELECT 1 FROM resenas WHERE id=?",(rid,)).fetchone():
        db.close(); return jsonify({"ok":False,"msg":"Resena no encontrada."}), 404
    if db.execute("SELECT 1 FROM votos WHERE resena_id=? AND session_key=?",(rid,sk)).fetchone():
        db.close(); return jsonify({"ok":False,"msg":"Ya votaste."})
    db.execute("INSERT INTO votos VALUES (?,?)",(rid,sk))
    col = "util_si" if voto=="si" else "util_no"
    db.execute(f"UPDATE resenas SET {col}={col}+1 WHERE id=?",(rid,))
    db.commit()
    row = db.execute("SELECT util_si,util_no FROM resenas WHERE id=?",(rid,)).fetchone()
    db.close()
    return jsonify({"ok":True,"util_si":row["util_si"],"util_no":row["util_no"]})

# ??????????????????????????????????????????????????????????????
#  ADMIN
# ??????????????????????????????????????????????????????????????

@app.route("/admin/login", methods=["GET","POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("user")==ADMIN_USER and request.form.get("password")==ADMIN_PASS:
            session["admin"] = True; return redirect(url_for("admin_dash"))
        flash("Credenciales incorrectas.")
    return render_template("admin_login.html")

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin",None); return redirect(url_for("admin_login"))


# Admin acciones
@app.route("/admin/cita/<int:cid>/estado",methods=["POST"])
@admin_required
def admin_cita_estado(cid):
    estado = request.form.get("estado")
    notas  = request.form.get("notas","")
    if estado in ["pendiente","confirmada","cancelada","completada"]:
        db = get_db()
        db.execute("UPDATE citas SET estado=?,notas_doctor=? WHERE id=?",(estado,notas,cid))
        row = db.execute("SELECT * FROM citas WHERE id=?",(cid,)).fetchone()
        if row and row["paciente_id"] and estado=="confirmada":
            db.execute("INSERT INTO notificaciones (paciente_id,tipo,titulo,mensaje) VALUES (?,?,?,?)",
                       (row["paciente_id"],"cita","Cita confirmada",f"Tu cita del {row['fecha']} a las {row['hora']} fue CONFIRMADA ✅"))
        db.commit(); db.close()
    return redirect(url_for("admin_dash")+"#citas")

@app.route("/admin/pago/<int:pid>/estado",methods=["POST"])
@admin_required
def admin_pago_estado(pid):
    estado = request.form.get("estado")
    if estado in ["pendiente","confirmado","rechazado"]:
        db = get_db()
        db.execute("UPDATE pagos SET estado=? WHERE id=?",(estado,pid))
        row = db.execute("SELECT * FROM pagos WHERE id=?",(pid,)).fetchone()
        if row and row["paciente_id"] and estado=="confirmado":
            db.execute("INSERT INTO notificaciones (paciente_id,tipo,titulo,mensaje) VALUES (?,?,?,?)",
                       (row["paciente_id"],"pago","Pago confirmado",f"Tu pago de S/ {row['monto']:.2f} fue CONFIRMADO ✅"))
            add_points(row["paciente_id"],10,"Pago confirmado")
        db.commit(); db.close()
    return redirect(url_for("admin_dash")+"#pagos")

@app.route("/admin/inventario/add",methods=["POST"])
@admin_required
def admin_inv_add():
    d = request.get_json()
    db = get_db()
    db.execute("INSERT INTO inventario (nombre,categoria,cantidad,unidad,minimo,precio_unit) VALUES (?,?,?,?,?,?)",
               (d.get("nombre"),d.get("categoria"),int(d.get("cantidad",0)),
                d.get("unidad"),int(d.get("minimo",5)),float(d.get("precio",0))))
    db.commit(); db.close()
    return jsonify({"ok":True})

@app.route("/admin/inventario/<int:iid>/update",methods=["POST"])
@admin_required
def admin_inv_update(iid):
    d = request.get_json()
    db = get_db()
    db.execute("UPDATE inventario SET cantidad=cantidad+? WHERE id=?",(int(d.get("delta",0)),iid))
    db.commit(); db.close()
    return jsonify({"ok":True})

@app.route("/admin/inventario/<int:iid>/delete",methods=["POST"])
@admin_required
def admin_inv_delete(iid):
    db = get_db()
    db.execute("DELETE FROM inventario WHERE id=?",(iid,))
    db.commit(); db.close()
    return redirect(url_for("admin_dash")+"#inventario")

def delete_paciente_relacionado(db, paciente_id):
    resena_ids = [r["id"] for r in db.execute("SELECT id FROM resenas WHERE paciente_id=?", (paciente_id,)).fetchall()]
    for rid in resena_ids:
        db.execute("DELETE FROM votos WHERE resena_id=?", (rid,))
    for table in ["notificaciones", "chat_mensajes", "historial_clinico", "escaneos", "citas", "pagos", "resenas"]:
        db.execute(f"DELETE FROM {table} WHERE paciente_id=?", (paciente_id,))
    db.execute("DELETE FROM pacientes WHERE id=?", (paciente_id,))


for tabla in ["escaneo","resena","paciente","cita","pago"]:
    @app.route(f"/admin/{tabla}/<int:rid>/delete",methods=["POST"],endpoint=f"del_{tabla}")
    @admin_required
    def gen_delete(rid, _tabla=tabla):
        map_t = {"escaneo":"escaneos","resena":"resenas","paciente":"pacientes","cita":"citas","pago":"pagos"}
        db = get_db()
        if _tabla == "paciente":
            delete_paciente_relacionado(db, rid)
        elif _tabla == "resena":
            db.execute("DELETE FROM votos WHERE resena_id=?", (rid,))
            db.execute("DELETE FROM resenas WHERE id=?", (rid,))
        else:
            db.execute(f"DELETE FROM {map_t[_tabla]} WHERE id=?",(rid,))
        db.commit(); db.close()
        return redirect(url_for("admin_dash")+f"#{map_t[_tabla]}")

# ?? HISTORIAL CL?NICO ??????????????????????????????????????????
@app.route("/admin/historial/<int:pac_id>")
@admin_required
def admin_historial(pac_id):
    db  = get_db()
    pac  = db.execute("SELECT * FROM pacientes WHERE id=?",(pac_id,)).fetchone()
    hist = db.execute("SELECT * FROM historial_clinico WHERE paciente_id=? ORDER BY fecha DESC",(pac_id,)).fetchall()
    esc  = db.execute("SELECT id,created_at,puntuacion,nivel_urgencia,diagnostico FROM escaneos WHERE paciente_id=? ORDER BY created_at DESC LIMIT 10",(pac_id,)).fetchall()
    cit  = db.execute("SELECT id,fecha,hora,servicio,estado FROM citas WHERE paciente_id=? ORDER BY fecha DESC LIMIT 10",(pac_id,)).fetchall()
    db.close()
    if not pac: return jsonify({"ok":False,"msg":"Paciente no encontrado"})
    return jsonify({"ok":True,
        "paciente":dict(pac), "edad":calcular_edad(pac["fecha_nac"]),
        "historial":[dict(h) for h in hist],
        "escaneos":[dict(e) for e in esc],
        "citas":[dict(c) for c in cit]})

@app.route("/admin/historial/add", methods=["POST"])
@admin_required
def admin_historial_add():
    d      = request.get_json()
    pac_id = d.get("paciente_id")
    if not pac_id: return jsonify({"ok":False,"msg":"Selecciona un paciente."})
    db = get_db()
    db.execute("""INSERT INTO historial_clinico
        (paciente_id,fecha,tipo,diagnostico,tratamiento_realizado,
         medicamentos,odontograma,observaciones,proxima_cita,creado_por)
        VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (pac_id, d.get("fecha") or date.today().isoformat(), d.get("tipo","Consulta"),
         d.get("diagnostico",""), d.get("tratamiento_realizado",""),
         d.get("medicamentos",""), d.get("odontograma",""),
         d.get("observaciones",""), d.get("proxima_cita",""),
         d.get("creado_por","Dr. Sergio Castillo")))
    db.execute("INSERT INTO notificaciones (paciente_id,tipo,titulo,mensaje) VALUES (?,?,?,?)",
               (pac_id,"historial","Nueva entrada en tu historial clínico",
                f"El Dr. Castillo agregó una entrada de {d.get('tipo','Consulta')} con fecha {d.get('fecha','')}. 📋"))
    db.commit(); db.close()
    return jsonify({"ok":True,"msg":"Historial clínico guardado."})

@app.route("/admin/historial/<int:hid>/delete", methods=["POST"])
@admin_required
def admin_historial_delete(hid):
    db = get_db()
    db.execute("DELETE FROM historial_clinico WHERE id=?",(hid,))
    db.commit(); db.close()
    return jsonify({"ok":True})

# ── BACKUP DB ──────────────────────────────────────────────────
@app.route("/admin/backup-db")
@admin_required
def backup_db():
    if not os.path.exists(DB):
        return "Base de datos no encontrada",404
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(os.path.abspath(DB), mimetype="application/x-sqlite3",
        as_attachment=True, download_name=f"coviden_backup_{ts}.db")

# ── VER BD ─────────────────────────────────────────────────────
@app.route("/admin/ver-bd")
@admin_required
def ver_bd():
    db = get_db()
    TABLAS = ["pacientes","escaneos","citas","pagos","historial_clinico",
              "notificaciones","resenas","inventario","chat_mensajes","votos"]
    data = {}
    for t in TABLAS:
        try:
            rows  = db.execute(f"SELECT * FROM {t} ORDER BY rowid DESC LIMIT 500").fetchall()
            cols  = list(rows[0].keys()) if rows else [c["name"] for c in db.execute(f"PRAGMA table_info({t})").fetchall()]
            count = db.execute(f"SELECT COUNT(*) as c FROM {t}").fetchone()["c"]
            data[t] = {"columns":cols,"rows":[dict(r) for r in rows],"count":count}
        except Exception as e:
            data[t] = {"columns":[],"rows":[],"count":0,"error":str(e)}
    db.close()
    db_size = os.path.getsize(DB) if os.path.exists(DB) else 0
    return render_template("ver_bd.html", data=data, db_size=db_size,
        now=datetime.now().strftime("%d/%m/%Y %H:%M"))

# ── ADMIN API KEYS ─────────────────────────────────────────────
@app.route("/admin/api-keys", methods=["POST"])
@admin_required
def admin_api_keys():
    updates = {}
    gemini_key = request.form.get("gemini_api_key", "").strip()
    groq_key = request.form.get("groq_api_key", "").strip()
    gemini_model_new = request.form.get("gemini_model", "").strip()
    groq_model_new = request.form.get("groq_model", "").strip()
    if gemini_key:
        updates["GEMINI_API_KEY"] = gemini_key
    if groq_key:
        updates["GROQ_API_KEY"] = groq_key
    if gemini_model_new:
        updates["GEMINI_MODEL"] = gemini_model_new
    if groq_model_new:
        updates["GROQ_MODEL"] = groq_model_new
    if updates:
        save_local_env(updates)
        os.environ.update(updates)
        refresh_ai_clients()
        session["admin_msg"] = "API keys guardadas en .env.local. Ese archivo no se sube a GitHub."
    else:
        session["admin_msg"] = "No se cambio nada. Escribe una key nueva o cambia un modelo."
    return redirect(url_for("admin_dash") + "#config")
# ── ADMIN DASHBOARD ────────────────────────────────────────────
@app.route("/admin")
@admin_required
def admin_dash():
    cumpleaneros_hoy = revisar_cumpleanos()
    db = get_db()
    t_pac    = db.execute("SELECT COUNT(*) as c FROM pacientes").fetchone()["c"]
    t_scan   = db.execute("SELECT COUNT(*) as c FROM escaneos").fetchone()["c"]
    t_citas  = db.execute("SELECT COUNT(*) as c FROM citas").fetchone()["c"]
    t_pago   = db.execute("SELECT COUNT(*) as c FROM pagos").fetchone()["c"]
    t_res    = db.execute("SELECT COUNT(*) as c FROM resenas").fetchone()["c"]
    t_hist   = db.execute("SELECT COUNT(*) as c FROM historial_clinico").fetchone()["c"]
    ingresos = db.execute("SELECT COALESCE(SUM(monto),0) as s FROM pagos WHERE estado='confirmado'").fetchone()["s"]
    pend_p   = db.execute("SELECT COUNT(*) as c FROM pagos WHERE estado='pendiente'").fetchone()["c"]
    pend_c   = db.execute("SELECT COUNT(*) as c FROM citas WHERE estado='pendiente'").fetchone()["c"]
    urg_alt  = db.execute("SELECT COUNT(*) as c FROM escaneos WHERE nivel_urgencia IN ('alta','urgente')").fetchone()["c"]
    avg_scan = db.execute("SELECT AVG(puntuacion) as a FROM escaneos").fetchone()["a"]
    avg_star = db.execute("SELECT AVG(estrellas) as a FROM resenas").fetchone()["a"]
    hoy_sc   = db.execute("SELECT COUNT(*) as c FROM escaneos WHERE DATE(created_at)=DATE('now','localtime')").fetchone()["c"]
    inv_bajo = db.execute("SELECT COUNT(*) as c FROM inventario WHERE cantidad <= minimo").fetchone()["c"]

    sc_dias  = db.execute("SELECT DATE(created_at) as d,COUNT(*) as c FROM escaneos WHERE created_at>=datetime('now','-14 days') GROUP BY d ORDER BY d").fetchall()
    urg_dist = db.execute("SELECT nivel_urgencia,COUNT(*) as c FROM escaneos GROUP BY nivel_urgencia").fetchall()
    pago_met = db.execute("SELECT metodo,COUNT(*) as c FROM pagos GROUP BY metodo").fetchall()
    pago_est = db.execute("SELECT estado,COUNT(*) as c FROM pagos GROUP BY estado").fetchall()
    star_d   = db.execute("SELECT estrellas,COUNT(*) as c FROM resenas GROUP BY estrellas ORDER BY estrellas").fetchall()
    pac_dias = db.execute("SELECT DATE(created_at) as d,COUNT(*) as c FROM pacientes WHERE created_at>=datetime('now','-14 days') GROUP BY d ORDER BY d").fetchall()
    cita_svc = db.execute("SELECT servicio,COUNT(*) as c FROM citas GROUP BY servicio ORDER BY c DESC LIMIT 6").fetchall()
    ingresos_dias = db.execute("SELECT DATE(created_at) as d,SUM(monto) as s FROM pagos WHERE estado='confirmado' AND created_at>=datetime('now','-14 days') GROUP BY d ORDER BY d").fetchall()
    expresion_dist = db.execute("SELECT expresion,COUNT(*) as c FROM escaneos WHERE expresion IS NOT NULL AND expresion!='' GROUP BY expresion ORDER BY c DESC LIMIT 8").fetchall()

    # Próximos cumpleaños (30 días)
    proximos_cumple = []
    hoy_d = date.today()
    pac_all = db.execute("SELECT id,nombre,apellido,fecha_nac,email,telefono FROM pacientes WHERE fecha_nac IS NOT NULL AND fecha_nac!=''").fetchall()
    for p in pac_all:
        try:
            _, mm, dd = p["fecha_nac"].split("-")
            prox = date(hoy_d.year, int(mm), int(dd))
            if prox < hoy_d: prox = date(hoy_d.year+1, int(mm), int(dd))
            dias = (prox - hoy_d).days
            if dias <= 30:
                proximos_cumple.append({"nombre":f"{p['nombre']} {p['apellido']}",
                    "fecha":prox.strftime("%d/%m"),"dias":dias,
                    "email":p["email"],"tel":p["telefono"]})
        except: continue
    proximos_cumple.sort(key=lambda x:x["dias"])

    pacientes  = db.execute("SELECT * FROM pacientes ORDER BY created_at DESC LIMIT 200").fetchall()
    escaneos   = db.execute("SELECT e.*,p.nombre as pn,p.apellido as pa FROM escaneos e LEFT JOIN pacientes p ON e.paciente_id=p.id ORDER BY e.created_at DESC LIMIT 80").fetchall()
    citas_all  = db.execute("SELECT c.*,p.nombre as pn FROM citas c LEFT JOIN pacientes p ON c.paciente_id=p.id ORDER BY c.fecha DESC,c.hora DESC LIMIT 80").fetchall()
    pagos_all  = db.execute("SELECT pg.*,p.nombre as pn FROM pagos pg LEFT JOIN pacientes p ON pg.paciente_id=p.id ORDER BY pg.created_at DESC LIMIT 80").fetchall()
    resenas    = db.execute("SELECT * FROM resenas ORDER BY created_at DESC LIMIT 50").fetchall()
    inventario = db.execute("SELECT * FROM inventario ORDER BY categoria,nombre").fetchall()
    db.close()

    return render_template("admin.html",
        t_pac=t_pac,t_scan=t_scan,t_citas=t_citas,t_pago=t_pago,t_res=t_res,t_hist=t_hist,
        ingresos=ingresos,pend_p=pend_p,pend_c=pend_c,urg_alt=urg_alt,
        avg_scan=round(avg_scan,1) if avg_scan else 0,
        avg_star=round(avg_star,1) if avg_star else 0,
        hoy=hoy_sc,inv_bajo=inv_bajo,
        cumpleaneros_hoy=cumpleaneros_hoy,
        proximos_cumple=proximos_cumple[:8],
        sc_dias=json.dumps([dict(r) for r in sc_dias]),
        urg_dist=json.dumps([dict(r) for r in urg_dist]),
        pago_met=json.dumps([dict(r) for r in pago_met]),
        pago_est=json.dumps([dict(r) for r in pago_est]),
        star_d=json.dumps([dict(r) for r in star_d]),
        pac_dias=json.dumps([dict(r) for r in pac_dias]),
        cita_svc=json.dumps([dict(r) for r in cita_svc]),
        ingresos_dias=json.dumps([dict(r) for r in ingresos_dias]),
        expresion_dist=json.dumps([dict(r) for r in expresion_dist]),
        pacientes=pacientes,escaneos=escaneos,citas_all=citas_all,
        pagos_all=pagos_all,resenas=resenas,inventario=inventario,
        gemini_model=GEMINI_MODEL,groq_model=GROQ_MODEL,
        gemini_key_status=mask_secret(GEMINI_API_KEY),
        groq_key_status=mask_secret(GROQ_API_KEY),
        admin_msg=session.pop("admin_msg", None),
        calcular_edad=calcular_edad)

# ── EXPORTES (PDF landscape celeste + Excel ampliado) ──────────
@app.route("/admin/export/<string:fmt>/<string:tipo>")
@admin_required
def export_data(fmt,tipo):
    db = get_db()
    SKY = (56,189,248)

    if fmt=="pdf":
        from fpdf import FPDF
        pdf = FPDF(orientation="L",unit="mm",format="A4")
        pdf.core_fonts_encoding = "cp1252"
        pdf.set_auto_page_break(True,15); pdf.add_page()
        # Header celeste
        pdf.set_fill_color(*SKY); pdf.rect(0,0,pdf.w,28,'F')
        pdf.set_font("Helvetica","B",18); pdf.set_text_color(0,0,0)
        pdf.set_xy(10,5); pdf.cell(0,10,"COVIDEN DENTAL",ln=True)
        pdf.set_font("Helvetica","",9); pdf.set_xy(10,17)
        pdf.cell(0,6,f"Reporte: {tipo.upper()}  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Av. Universitaria 8156, Comas, Lima",ln=True)
        pdf.set_text_color(0,0,0); pdf.ln(10)
        pw = pdf.w-20

        def hdr(cols):
            pdf.set_font("Helvetica","B",8); pdf.set_fill_color(224,242,254); pdf.set_text_color(0,0,0)
            for h,w in cols: pdf.cell(w,8,h,border=1,align="C",fill=True)
            pdf.ln(); pdf.set_font("Helvetica","",7.5)

        if tipo=="escaneos":
            rows=db.execute("""SELECT e.*,p.nombre pn,p.apellido pa,p.email pe,p.telefono pt,p.fecha_nac pfn,p.sexo psx
                FROM escaneos e LEFT JOIN pacientes p ON e.paciente_id=p.id ORDER BY e.created_at DESC""").fetchall()
            pdf.set_font("Helvetica","B",11);pdf.cell(0,7,"ESCANEOS DENTALES — Análisis con IA",ln=True);pdf.ln(2)
            hdr([("ID",10),("Paciente",44),("Email",50),("Teléfono",26),("Edad",14),("Sexo",14),("Score",16),("Urgencia",22),("Expresión",26),("Fecha",36)])
            for r in rows:
                nom=(f"{r['pn']} {r['pa']}" if r["pn"] else r["nombre_libre"] or "—")
                edad=calcular_edad(r["pfn"])
                pdf.cell(10,7,str(r["id"]),border=1,align="C")
                pdf.cell(44,7,nom[:26],border=1)
                pdf.cell(50,7,(r["pe"] or r["email_libre"] or "—")[:30],border=1)
                pdf.cell(26,7,(r["pt"] or r["telefono_libre"] or "—")[:14],border=1,align="C")
                pdf.cell(14,7,str(edad) if edad else "—",border=1,align="C")
                pdf.cell(14,7,(r["psx"] or "—")[:3],border=1,align="C")
                pdf.cell(16,7,f"{r['puntuacion'] or 0}/10",border=1,align="C")
                pdf.cell(22,7,(r["nivel_urgencia"] or "—").upper(),border=1,align="C")
                pdf.cell(26,7,(r["expresion"] or "—")[:16],border=1,align="C")
                pdf.cell(36,7,(r["created_at"] or "")[:16],border=1,align="C")
                pdf.ln()

        elif tipo=="pagos":
            rows=db.execute("""SELECT pg.*,p.nombre pn,p.apellido pa FROM pagos pg
                LEFT JOIN pacientes p ON pg.paciente_id=p.id ORDER BY pg.created_at DESC""").fetchall()
            pdf.set_font("Helvetica","B",11);pdf.cell(0,7,"PAGOS Y RESERVAS",ln=True);pdf.ln(2)
            hdr([("ID",10),("Paciente",40),("Email",46),("Teléfono",26),("Servicio",44),("Monto",20),("Método",22),("Estado",24),("Referencia",26),("Fecha",26)])
            total=0
            for r in rows:
                nom=(f"{r['pn']} {r['pa']}" if r["pn"] else r["nombre_libre"] or "—")
                pdf.cell(10,7,str(r["id"]),border=1,align="C")
                pdf.cell(40,7,nom[:22],border=1)
                pdf.cell(46,7,(r["email_libre"] or "—")[:27],border=1)
                pdf.cell(26,7,(r["tel_libre"] or "—")[:14],border=1,align="C")
                pdf.cell(44,7,(r["servicio"] or "")[:25],border=1)
                pdf.cell(20,7,f"S/{r['monto']:.2f}",border=1,align="R")
                pdf.cell(22,7,(r["metodo"] or "").upper()[:10],border=1,align="C")
                pdf.cell(24,7,(r["estado"] or "").upper()[:12],border=1,align="C")
                pdf.cell(26,7,(r["referencia"] or "—")[:14],border=1,align="C")
                pdf.cell(26,7,(r["created_at"] or "")[:10],border=1,align="C")
                pdf.ln()
                if r["estado"]=="confirmado": total+=r["monto"] or 0
            pdf.ln(3);pdf.set_font("Helvetica","B",11)
            pdf.cell(0,8,f"TOTAL INGRESOS CONFIRMADOS: S/ {total:.2f}",ln=True,align="R")

        elif tipo=="citas":
            rows=db.execute("""SELECT c.*,p.nombre pn,p.apellido pa,p.email pe FROM citas c
                LEFT JOIN pacientes p ON c.paciente_id=p.id ORDER BY c.fecha DESC,c.hora DESC""").fetchall()
            pdf.set_font("Helvetica","B",11);pdf.cell(0,7,"AGENDA DE CITAS",ln=True);pdf.ln(2)
            hdr([("ID",10),("Paciente",44),("Email",48),("Teléfono",26),("Servicio",42),("Fecha",22),("Hora",16),("Estado",24),("Motivo",36)])
            for r in rows:
                nom=(f"{r['pn']} {r['pa']}" if r["pn"] else r["nombre_libre"] or "—")
                pdf.cell(10,7,str(r["id"]),border=1,align="C")
                pdf.cell(44,7,nom[:25],border=1)
                pdf.cell(48,7,(r["pe"] or r["email_libre"] or "—")[:28],border=1)
                pdf.cell(26,7,(r["tel_libre"] or "—")[:14],border=1,align="C")
                pdf.cell(42,7,(r["servicio"] or "")[:24],border=1)
                pdf.cell(22,7,(r["fecha"] or ""),border=1,align="C")
                pdf.cell(16,7,(r["hora"] or ""),border=1,align="C")
                pdf.cell(24,7,(r["estado"] or "").upper(),border=1,align="C")
                pdf.cell(36,7,(r["motivo"] or "—")[:20],border=1)
                pdf.ln()

        elif tipo=="pacientes":
            rows=db.execute("SELECT * FROM pacientes ORDER BY created_at DESC").fetchall()
            pdf.set_font("Helvetica","B",11);pdf.cell(0,7,"PACIENTES REGISTRADOS",ln=True);pdf.ln(2)
            hdr([("ID",10),("Nombre",42),("Email",50),("Teléfono",26),("DNI",22),("Edad",14),("Sexo",14),("Nivel",20),("Puntos",18),("Registro",30)])
            for r in rows:
                edad=calcular_edad(r["fecha_nac"])
                pdf.cell(10,7,str(r["id"]),border=1,align="C")
                pdf.cell(42,7,f"{r['nombre']} {r['apellido']}"[:24],border=1)
                pdf.cell(50,7,(r["email"] or "")[:28],border=1)
                pdf.cell(26,7,(r["telefono"] or "—")[:14],border=1,align="C")
                pdf.cell(22,7,(r["dni"] or "—"),border=1,align="C")
                pdf.cell(14,7,str(edad) if edad else "—",border=1,align="C")
                pdf.cell(14,7,(r["sexo"] or "—")[:3],border=1,align="C")
                pdf.cell(20,7,(r["nivel"] or "Bronce"),border=1,align="C")
                pdf.cell(18,7,str(r["puntos"] or 0),border=1,align="C")
                pdf.cell(30,7,(r["created_at"] or "")[:10],border=1,align="C")
                pdf.ln()

        elif tipo=="historial":
            pac_id=request.args.get("paciente_id",type=int)
            if pac_id:
                pac=db.execute("SELECT * FROM pacientes WHERE id=?",(pac_id,)).fetchone()
                rows=db.execute("SELECT * FROM historial_clinico WHERE paciente_id=? ORDER BY fecha DESC",(pac_id,)).fetchall()
                if pac:
                    edad=calcular_edad(pac["fecha_nac"])
                    pdf.set_font("Helvetica","B",12)
                    pdf.cell(0,8,f"HISTORIAL CLÍNICO: {pac['nombre']} {pac['apellido']}",ln=True)
                    pdf.set_font("Helvetica","",9)
                    pdf.cell(0,6,f"Email: {pac['email']}  |  Tel: {pac['telefono'] or '—'}  |  DNI: {pac['dni'] or '—'}  |  Edad: {edad or '—'} años  |  Sexo: {pac['sexo'] or '—'}",ln=True)
                    if pac["alergias"]: pdf.set_font("Helvetica","B",9);pdf.cell(0,6,f"⚠ Alergias: {pac['alergias']}",ln=True)
                    pdf.ln(4)
                hdr([("Fecha",22),("Tipo",28),("Diagnóstico",64),("Tratamiento",64),("Medicamentos",44),("Observaciones",44),("Próx. Cita",24)])
                for r in rows:
                    pdf.cell(22,7,(r["fecha"] or "")[:10],border=1,align="C")
                    pdf.cell(28,7,(r["tipo"] or "")[:16],border=1)
                    pdf.cell(64,7,(r["diagnostico"] or "—")[:38],border=1)
                    pdf.cell(64,7,(r["tratamiento_realizado"] or "—")[:38],border=1)
                    pdf.cell(44,7,(r["medicamentos"] or "—")[:26],border=1)
                    pdf.cell(44,7,(r["observaciones"] or "—")[:26],border=1)
                    pdf.cell(24,7,(r["proxima_cita"] or "—")[:12],border=1,align="C")
                    pdf.ln()

        elif tipo=="resenas":
            rows=db.execute("SELECT * FROM resenas ORDER BY created_at DESC").fetchall()
            pdf.set_font("Helvetica","B",11);pdf.cell(0,7,"RESEÑAS DE PACIENTES",ln=True);pdf.ln(2)
            hdr([("ID",10),("Nombre",46),("Servicio",50),("Estrellas",22),("Comentario",120),("Fecha",30)])
            for r in rows:
                pdf.cell(10,7,str(r["id"]),border=1,align="C")
                pdf.cell(46,7,(r["nombre"] or "")[:26],border=1)
                pdf.cell(50,7,(r["servicio"] or "")[:28],border=1)
                pdf.cell(22,7,f"{r["estrellas"]}/5",border=1,align="C")
                pdf.cell(120,7,(r["comentario"] or "—")[:70],border=1)
                pdf.cell(30,7,(r["created_at"] or "")[:10],border=1,align="C")
                pdf.ln()

        db.close()
        buf=io.BytesIO(pdf.output());buf.seek(0)
        return send_file(buf,mimetype="application/pdf",as_attachment=True,
            download_name=f"coviden_{tipo}_{datetime.now().strftime('%Y%m%d')}.pdf")

    elif fmt=="excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
        wb=Workbook();ws=wb.active
        hf=Font(bold=True,color="000000",size=11)
        hfl=PatternFill("solid",fgColor="38BDF8")  # celeste sky-500, texto negro
        thin=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
        def headers(cols):
            ws.append(cols)
            for cell in ws[1]:
                cell.font=hf;cell.fill=hfl
                cell.alignment=Alignment(horizontal="center",vertical="center");cell.border=thin
            ws.row_dimensions[1].height=22

        if tipo=="escaneos":
            ws.title="Escaneos"
            headers(["ID","Paciente","Email","Teléfono","Edad","Sexo","Puntuación","Urgencia","Expresión","Diagnóstico","Mensaje IA","Fecha"])
            for r in db.execute("""SELECT e.*,p.nombre pn,p.apellido pa,p.email pe,p.telefono pt,p.fecha_nac pfn,p.sexo psx
                FROM escaneos e LEFT JOIN pacientes p ON e.paciente_id=p.id ORDER BY e.created_at DESC""").fetchall():
                nom=f"{r['pn']} {r['pa']}" if r["pn"] else (r["nombre_libre"] or "—")
                edad=calcular_edad(r["pfn"])
                ws.append([r["id"],nom,r["pe"] or r["email_libre"],r["pt"] or r["telefono_libre"],
                    edad,r["psx"],r["puntuacion"],r["nivel_urgencia"],r["expresion"],r["diagnostico"],r["mensaje"],r["created_at"]])

        elif tipo=="pagos":
            ws.title="Pagos"
            headers(["ID","Paciente","Email","Teléfono","Servicio","Monto S/","Método","Estado","Referencia","Notas","Fecha"])
            total=0
            for r in db.execute("SELECT pg.*,p.nombre pn,p.apellido pa FROM pagos pg LEFT JOIN pacientes p ON pg.paciente_id=p.id ORDER BY pg.created_at DESC").fetchall():
                nom=f"{r['pn']} {r['pa']}" if r["pn"] else (r["nombre_libre"] or "—")
                ws.append([r["id"],nom,r["email_libre"],r["tel_libre"],r["servicio"],r["monto"],r["metodo"],r["estado"],r["referencia"],r["notas"],r["created_at"]])
                if r["estado"]=="confirmado": total+=r["monto"] or 0
            ws.append([]);ws.append(["","","","","TOTAL CONFIRMADO:",total])

        elif tipo=="citas":
            ws.title="Citas"
            headers(["ID","Paciente","Email","Teléfono","Servicio","Fecha","Hora","Estado","Motivo","Notas Doctor","Creado"])
            for r in db.execute("SELECT c.*,p.nombre pn,p.apellido pa,p.email pe FROM citas c LEFT JOIN pacientes p ON c.paciente_id=p.id ORDER BY c.fecha DESC").fetchall():
                nom=f"{r['pn']} {r['pa']}" if r["pn"] else (r["nombre_libre"] or "—")
                ws.append([r["id"],nom,r["pe"] or r["email_libre"],r["tel_libre"],r["servicio"],r["fecha"],r["hora"],r["estado"],r["motivo"],r["notas_doctor"],r["created_at"]])

        elif tipo=="pacientes":
            ws.title="Pacientes"
            headers(["ID","Nombre","Apellido","Email","Teléfono","DNI","Fecha Nac","Edad","Sexo","Dirección","Alergias","Nivel","Puntos","Registro"])
            for r in db.execute("SELECT * FROM pacientes ORDER BY created_at DESC").fetchall():
                ws.append([r["id"],r["nombre"],r["apellido"],r["email"],r["telefono"],r["dni"],r["fecha_nac"],
                    calcular_edad(r["fecha_nac"]),r["sexo"],r["direccion"],r["alergias"],r["nivel"],r["puntos"],r["created_at"]])

        elif tipo=="historial":
            ws.title="HistorialClinico"
            headers(["ID","Paciente","Email","Edad","Sexo","Alergias","Fecha","Tipo","Diagnóstico","Tratamiento Realizado","Medicamentos","Odontograma","Observaciones","Próxima Cita","Creado Por"])
            for r in db.execute("""SELECT h.*,p.nombre pn,p.apellido pa,p.email pe,p.fecha_nac pfn,p.sexo psx,p.alergias pal
                FROM historial_clinico h LEFT JOIN pacientes p ON h.paciente_id=p.id ORDER BY h.fecha DESC""").fetchall():
                nom=f"{r['pn']} {r['pa']}" if r["pn"] else "—"
                ws.append([r["id"],nom,r["pe"],calcular_edad(r["pfn"]),r["psx"],r["pal"],
                    r["fecha"],r["tipo"],r["diagnostico"],r["tratamiento_realizado"],
                    r["medicamentos"],r["odontograma"],r["observaciones"],r["proxima_cita"],r["creado_por"]])

        elif tipo=="resenas":
            ws.title="Resenas"
            headers(["ID","Nombre","Servicio","Estrellas","Comentario","Votos Útil","Votos No","Fecha"])
            for r in db.execute("SELECT * FROM resenas ORDER BY created_at DESC").fetchall():
                ws.append([r["id"],r["nombre"],r["servicio"],r["estrellas"],r["comentario"],r["util_si"],r["util_no"],r["created_at"]])

        for col in ws.columns:
            ml=max((len(str(c.value or "")) for c in col),default=10)
            ws.column_dimensions[col[0].column_letter].width=min(ml+4,55)
        db.close()
        buf=io.BytesIO();wb.save(buf);buf.seek(0)
        return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,download_name=f"coviden_{tipo}_{datetime.now().strftime('%Y%m%d')}.xlsx")

    db.close()
    return "Formato no soportado",400
# ── JINJA FILTER ───────────────────────────────────────────────
@app.template_filter("from_json")
def from_json_filter(s):
    try: return json.loads(s or "[]")
    except: return []

# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    init_db()
    print("="*60)
    print("  COVIDEN DENTAL v5.0 PREMIUM")
    print(f"  Gemini: {GEMINI_MODEL}")
    print(f"  Groq:   {GROQ_MODEL}")
    print("="*60)
    print("  Sitio     -> http://localhost:5000")
    print("  Mi cuenta -> http://localhost:5000/mi-cuenta")
    print("  Admin     -> http://localhost:5000/admin  (admin/coviden2024)")
    print("  Ver BD    -> http://localhost:5000/admin/ver-bd")
    print("  Backup    -> http://localhost:5000/admin/backup-db")
    print("="*60)
    app.run(debug=True, port=5000, threaded=True)




